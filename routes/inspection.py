"""点检模板与点检记录相关路由"""
import os
from flask import Blueprint, jsonify, request, make_response, send_from_directory
from services.inspection_service import (
    get_templates, create_template, update_template,
    delete_template, copy_templates, import_templates,
    create_inspection, list_inspections,
    get_inspection_by_id, delete_inspection,
    check_today_inspected,
    save_inspection_photo, get_inspection_photo_path,
)
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

inspection_bp = Blueprint("inspection", __name__)


# ---- 点检模板 ----

@inspection_bp.route("/api/inspection-templates", methods=["GET"])
def query_templates():
    try:
        machine_id = request.args.get("machine_id", type=int)
        if not machine_id:
            return jsonify({"error": "缺少 machine_id 参数"}), 400
        return jsonify(get_templates(machine_id))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection-templates", methods=["POST"])
def add_template():
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "请求体必须为 JSON"}), 400
        new_id = create_template(data)
        return jsonify({"success": True, "id": new_id}), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection-templates/<int:template_id>", methods=["PUT"])
def edit_template(template_id):
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "请求体必须为 JSON"}), 400
        ok = update_template(template_id, data)
        if not ok:
            return jsonify({"error": "更新失败"}), 404
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection-templates/<int:template_id>", methods=["DELETE"])
def remove_template(template_id):
    try:
        ok = delete_template(template_id)
        if not ok:
            return jsonify({"error": "记录不存在"}), 404
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection-templates/copy", methods=["POST"])
def copy_template():
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "请求体必须为 JSON"}), 400
        source_id = data.get("source_machine_id")
        target_id = data.get("target_machine_id")
        if not source_id or not target_id:
            return jsonify({"error": "缺少 source_machine_id 或 target_machine_id"}), 400
        count = copy_templates(source_id, target_id)
        return jsonify({"success": True, "copied": count})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection-templates/import", methods=["POST"])
def import_template():
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "请求体必须为 JSON"}), 400
        machine_id = data.get("machine_id")
        items = data.get("items", [])
        if not machine_id:
            return jsonify({"error": "缺少 machine_id"}), 400
        if not items:
            return jsonify({"error": "没有可导入的项目"}), 400
        count = import_templates(machine_id, items)
        return jsonify({"success": True, "imported": count})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---- 点检记录 ----

@inspection_bp.route("/api/inspection", methods=["POST"])
def add_inspection():
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "请求体必须为 JSON"}), 400
        mid = data.get("machine_id")
        if mid and check_today_inspected(int(mid)):
            return jsonify({"error": "该设备今日已点检，不能重复提交。如需重新填写，请联系管理员删除今日记录后重试"}), 400
        new_id = create_inspection(data)
        return jsonify({"success": True, "id": new_id}), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection", methods=["GET"])
def query_inspections():
    try:
        machine_id = request.args.get("machine_id", type=int)
        date_from = request.args.get("from")
        date_to = request.args.get("to")
        return jsonify(list_inspections(
            machine_id=machine_id,
            date_from=date_from,
            date_to=date_to,
        ))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection/<int:record_id>", methods=["GET"])
def get_inspection(record_id):
    try:
        r = get_inspection_by_id(record_id)
        if r is None:
            return jsonify({"error": "记录不存在"}), 404
        return jsonify(r)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection/<int:record_id>", methods=["DELETE"])
def remove_inspection(record_id):
    try:
        ok = delete_inspection(record_id)
        if not ok:
            return jsonify({"error": "记录不存在"}), 404
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---- 点检照片 ----

@inspection_bp.route("/api/inspection/upload-photo", methods=["POST"])
def upload_photo():
    """上传点检照片"""
    try:
        if "file" not in request.files:
            return jsonify({"error": "没有上传文件"}), 400
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "未选择文件"}), 400
        machine_id = request.form.get("machine_id", type=int)
        if not machine_id:
            return jsonify({"error": "缺少 machine_id"}), 400
        filename = save_inspection_photo(machine_id, file)
        return jsonify({"success": True, "filename": filename})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@inspection_bp.route("/api/inspection-photos/<filename>", methods=["GET"])
def get_photo(filename):
    """获取点检照片"""
    path = get_inspection_photo_path(filename)
    if not path:
        return jsonify({"error": "照片不存在"}), 404
    from config import UPLOAD_DIR
    photo_dir = os.path.join(UPLOAD_DIR, "inspection_photos")
    return send_from_directory(photo_dir, filename)


# ---- 导出 Excel ----

@inspection_bp.route("/api/inspection/export", methods=["GET"])
def export_inspections():
    try:
        machine_id = request.args.get("machine_id", type=int)
        records = list_inspections(machine_id=machine_id, limit=1000)
        if not records:
            return jsonify({"error": "没有可导出的点检记录"}), 400

        grouped = {}
        for r in records:
            key = r.get("machine_code", "未知机床")
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(r)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for machine_code, recs in grouped.items():
            sheet_name = machine_code[:31]
            ws = wb.create_sheet(title=sheet_name)

            all_items = []
            if recs[0].get("details"):
                for d in recs[0]["details"]:
                    all_items.append(d.get("item_name", ""))

            ws.cell(row=1, column=1, value="点检人")
            ws.cell(row=2, column=1, value="日期")
            for i, item_name in enumerate(all_items):
                ws.cell(row=i + 3, column=1, value=item_name)

            for col_idx, rec in enumerate(recs):
                col = col_idx + 2
                ws.cell(row=1, column=col, value=rec.get("operator_name", ""))
                ws.cell(row=2, column=col, value=rec.get("created_at", ""))
                details = rec.get("details", [])
                for i, d in enumerate(details):
                    result = d.get("result", "")
                    note = d.get("note", "")
                    if note:
                        result = result + "（" + note + "）"
                    ws.cell(row=i + 3, column=col, value=result)

            ws.column_dimensions['A'].width = 45
            for ci in range(len(recs)):
                ws.column_dimensions[get_column_letter(ci + 2)].width = 25

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        resp = make_response(output.read())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"] = "attachment; filename=inspection_records.xlsx"
        return resp
    except Exception as e:
        return jsonify({"error": str(e)}), 500
